import shutil
from pathlib import Path
from openpyxl import load_workbook
import streamlit as st
import json
import os

# Configuração da página
st.set_page_config(
    page_title="Estrutura do Laudo Pericial",
    page_icon="📋",
    layout="wide"
)

# Função para carregar orientações existentes


def carregar_orientacoes():
    """Carrega as orientações salvas no arquivo JSON"""
    try:
        if os.path.exists("orientacoes_estrutura_laudo.json"):
            with open("orientacoes_estrutura_laudo.json", "r", encoding="utf-8") as f:
                return json.load(f)
    except Exception as e:
        st.error(f"❌ Erro ao carregar orientações: {str(e)}")
    return None


# Carregar orientações existentes ao iniciar a página
orientacoes_carregadas = carregar_orientacoes()

# Título principal
st.markdown("# 📋 Estrutura do Laudo Pericial")
st.markdown("### Orientações para Geração do Relatório Final")

st.markdown("---")

# Informações sobre o uso
st.info("""
**Como usar esta página:**
- Preencha cada campo com as orientações específicas para a IA
- A IA utilizará estas orientações como base para gerar o relatório final
- Os campos podem ser editados conforme necessário para cada tipo de laudo
- As orientações são salvas automaticamente e carregadas quando a página é aberta
""")

# Campo Histórico
st.markdown("## 📖 Histórico")
st.markdown("**Descrição:** Orientações para a seção de histórico do laudo")
historico_orientacoes = st.text_area(
    "Orientações para o Histórico:",
    value=orientacoes_carregadas.get(
        "historico", "") if orientacoes_carregadas else "",
    height=150,
    help="Instruções para a IA sobre como estruturar a seção de histórico"
)

# Campo Objetivo
st.markdown("## 🎯 Objetivo")
st.markdown("**Descrição:** Orientações para a seção de objetivo da perícia")
objetivo_orientacoes = st.text_area(
    "Orientações para o Objetivo:",
    value=orientacoes_carregadas.get(
        "objetivo", "") if orientacoes_carregadas else "",
    height=150,
    help="Instruções para a IA sobre como estruturar a seção de objetivo"
)

# Campo Quesitos
st.markdown("## ❓ Quesitos")
st.markdown("**Descrição:** Orientações para a seção de quesitos da perícia")
quesitos_orientacoes = st.text_area(
    "Orientações para os Quesitos:",
    value=orientacoes_carregadas.get(
        "quesitos", "") if orientacoes_carregadas else "",
    height=150,
    help="Instruções para a IA sobre como tratar os quesitos da perícia"
)


# Campo Conclusão/Resposta aos Quesitos
st.markdown("## ✅ Conclusão/Resposta aos Quesitos")
st.markdown("**Descrição:** Orientações para a seção de conclusão e respostas")
conclusao_orientacoes = st.text_area(
    "Orientações para a Conclusão:",
    value=orientacoes_carregadas.get(
        "conclusao", "") if orientacoes_carregadas else "",
    height=150,
    help="Instruções para a IA sobre como estruturar a seção de conclusão"
)

st.markdown("---")

# Seção de salvamento das orientações
st.markdown("## 💾 Salvar Orientações")

# Botão para salvar as orientações
if st.button("💾 Salvar Orientações da Estrutura", type="primary"):
    # Criar dicionário com todas as orientações
    orientacoes = {
        "historico": historico_orientacoes,
        "objetivo": objetivo_orientacoes,
        "quesitos": quesitos_orientacoes,
        "conclusao": conclusao_orientacoes,
        "data_criacao": str(st.session_state.get("data_atual", "N/A")),
        "versao": "1.0"
    }

    # Salvar em arquivo JSON
    try:
        with open("orientacoes_estrutura_laudo.json", "w", encoding="utf-8") as f:
            json.dump(orientacoes, f, ensure_ascii=False, indent=2)

        st.success("✅ Orientações salvas com sucesso!")
        st.info("As orientações foram salvas no arquivo 'orientacoes_estrutura_laudo.json' e podem ser utilizadas pela página 05_GeraRelatorio")

        # Recarregar a página para mostrar as orientações salvas
        st.rerun()

    except Exception as e:
        st.error(f"❌ Erro ao salvar as orientações: {str(e)}")

# Botão para recarregar orientações existentes
if st.button("📂 Recarregar Orientações do Arquivo", type="secondary"):
    try:
        if os.path.exists("orientacoes_estrutura_laudo.json"):
            orientacoes_carregadas = carregar_orientacoes()
            if orientacoes_carregadas:
                st.success("✅ Orientações recarregadas com sucesso!")
                st.info("As orientações foram recarregadas do arquivo existente")
                st.rerun()  # Recarregar a página para mostrar as orientações
            else:
                st.warning("⚠️ Erro ao carregar as orientações")
        else:
            st.warning("⚠️ Nenhum arquivo de orientações encontrado")

    except Exception as e:
        st.error(f"❌ Erro ao carregar as orientações: {str(e)}")

# Mostrar status das orientações
if orientacoes_carregadas:
    st.success("✅ Orientações carregadas do arquivo existente")
    st.info(
        f"Versão: {orientacoes_carregadas.get('versao', 'N/A')} | Data: {orientacoes_carregadas.get('data_criacao', 'N/A')}")
else:
    st.info("ℹ️ Nenhuma orientação carregada. Preencha os campos e salve para criar o arquivo de orientações.")

# Informações adicionais
st.markdown("---")
st.markdown("## 📝 Informações Importantes")

st.markdown("""
**Como a página 05_GeraRelatorio utilizará estas orientações:**

1. **Histórico**: A IA incluirá as informações conforme as orientações definidas
2. **Objetivo**: Será usado para estruturar a seção de objetivos da perícia
3. **Quesitos**: A IA copiará exatamente o conteúdo da página 01_CriaRelatorio
4. **Isolamento e Preservação**: Orientações para descrever o estado do local
5. **Exames**: Diretrizes para detalhar os procedimentos técnicos realizados
6. **Conclusão**: Instruções para responder aos quesitos de forma estruturada

**Dica**: Estas orientações funcionam como um "prompt" para a IA, fornecendo o contexto necessário para gerar um laudo pericial completo e bem estruturado.
""")

# Rodapé
st.markdown("---")
st.markdown(
    "*Página de Estrutura do Laudo Pericial - Sistema de Geração de Relatórios*")


def limpar_historico_excel(
    caminho_arquivo="historico_relatorios.xlsx",
    nome_planilha=None,           # None = todas as planilhas
    manter_cabecalho=True,        # True = mantém a linha 1
    criar_backup=True             # True = cria backup .bak ao lado
):
    caminho = Path(caminho_arquivo)
    if not caminho.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {caminho}")

    if criar_backup:
        backup_path = caminho.with_suffix(caminho.suffix + ".bak")
        shutil.copy2(caminho, backup_path)

    wb = load_workbook(caminho)
    planilhas = [wb[nome_planilha]] if nome_planilha else wb.worksheets

    for ws in planilhas:
        max_row = ws.max_row or 0
        if max_row <= 1:
            continue  # já só com cabeçalho (ou vazia)
        if manter_cabecalho:
            # Remove a partir da 2ª linha até o final
            ws.delete_rows(2, max_row - 1)
        else:
            # Remove tudo
            ws.delete_rows(1, max_row)

    wb.save(caminho)
    return True


if st.button("Limpar histórico (manter cabeçalho)"):
    try:
        limpar_historico_excel("historico_relatorios.xlsx")
        st.success("Histórico limpo com sucesso (cabeçalho preservado).")
    except Exception as e:
        st.error(f"Falha ao limpar: {e}")
